from __future__ import annotations

import argparse
import hashlib
from collections import defaultdict
from datetime import UTC, datetime
from pathlib import Path

from openpyxl import load_workbook

from backend.models.evidence import ConfidenceLevel, EvidenceRecord, EvidenceType
from backend.models.geography import GeographicIdentity

from .store import EvidenceStore

EXPECTED_SHA256 = "3fb48f43fe4b3a500371e46471af8f3da8a2736b07578a0d3e21731c04cc5142"
SOURCE_URL = (
    "https://dahd.gov.in/sites/default/files/2023-07/VillageAndWardLevelDataMale-Female.xlsx"
)
SPECIES = ("cattle", "buffalo", "sheep", "goat", "pig")


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _geo_id(kind: str, district: str, parent: str, locality: str) -> str:
    identity = "|".join((kind, district, parent, locality)).casefold()
    digest = hashlib.sha1(identity.encode(), usedforsecurity=False).hexdigest()[:12]
    return f"DS057:WB:{kind}:{digest}"


def ingest_livestock_workbook(
    workbook_path: str | Path,
    store: EvidenceStore,
    *,
    verify_checksum: bool = True,
    expected_sha256: str | None = EXPECTED_SHA256,
    maximum_source_rows: int | None = None,
) -> dict[str, int]:
    path = Path(workbook_path)
    actual_hash = _sha256(path)
    if verify_checksum and expected_sha256 is None:
        raise ValueError("an expected SHA-256 is required when verification is enabled")
    if verify_checksum and actual_hash != expected_sha256:
        raise ValueError(f"regional workbook SHA-256 mismatch: {actual_hash}")
    workbook = load_workbook(path, read_only=False, data_only=True)
    aggregates: dict[tuple[str, str, str, str], dict[str, dict[str, int]]] = defaultdict(
        lambda: defaultdict(lambda: defaultdict(int))
    )
    source_references: dict[tuple[str, str, str, str], dict[str, list[int]]] = defaultdict(
        lambda: defaultdict(list)
    )
    source_rows = 0
    for sheet in workbook.worksheets:
        if sheet.title == "README":
            continue
        is_rural = sheet.title.startswith("Rural")
        sex = "male" if "Male" in sheet.title else "female"
        headers = [str(cell.value) for cell in sheet[1]]
        positions = {name: index for index, name in enumerate(headers)}
        source_row_position = positions.get("source_row")
        parent_column = "block_name" if is_rural else "town_name"
        locality_column = "village_name" if is_rural else "ward_name"
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if maximum_source_rows is not None and source_rows >= maximum_source_rows:
                break
            district = str(row[positions["district_name"]] or "").strip()
            parent = str(row[positions[parent_column]] or "").strip()
            locality = str(row[positions[locality_column]] or "").strip()
            if not district or not locality:
                continue
            kind = "RURAL" if is_rural else "URBAN"
            key = (kind, district, parent, locality)
            for species in SPECIES:
                aggregates[key][species][sex] += int(row[positions[species]] or 0)
            if source_row_position is not None and row[source_row_position] is not None:
                source_references[key][sex].append(int(row[source_row_position]))
            source_rows += 1
        if maximum_source_rows is not None and source_rows >= maximum_source_rows:
            break

    retrieved_at = datetime.fromtimestamp(path.stat().st_mtime, UTC)
    evidence_count = 0
    with store.transaction():
        for (kind, district, parent, locality), species_data in aggregates.items():
            geo_id = _geo_id(kind, district, parent, locality)
            store.put_geography(
                GeographicIdentity(
                    geo_id=geo_id,
                    district=district,
                    locality=locality,
                    locality_type="VILLAGE" if kind == "RURAL" else "WARD",
                    block=parent if kind == "RURAL" else None,
                    municipality=parent if kind == "URBAN" else None,
                    source_ids=["DS057"],
                    quality_flags=[
                        "DATASET_LABEL_ID_NOT_OFFICIAL_LGD_OR_CENSUS_CODE",
                        "DISTRICT_SPELLING_PRESERVED_FROM_SOURCE",
                    ],
                )
            )
            for species, breakdown in species_data.items():
                total = sum(breakdown.values())
                store.put_evidence(
                    EvidenceRecord(
                        id=f"DS057:{geo_id}:{species}",
                        variable=f"livestock_population_{species}",
                        value=total,
                        unit="animals",
                        geography=locality,
                        geo_id=geo_id,
                        source_id="DS057",
                        source_url=SOURCE_URL,
                        source_dataset="20th Livestock Census village and ward data",
                        observation_date=None,
                        retrieved_at=retrieved_at,
                        evidence_type=EvidenceType.OBSERVED,
                        confidence=ConfidenceLevel.HIGH,
                        quality_flags=["MALE_AND_FEMALE_ROWS_SUMMED"],
                        methodology_version="ds057-regional-aggregate-v1",
                        raw_reference=path.name,
                        attributes={
                            "sex_breakdown": dict(breakdown),
                            "source_rows": dict(
                                source_references[(kind, district, parent, locality)]
                            ),
                            "census_reference_year": 2019,
                            "dataset_version": "20th Livestock Census 2019",
                        },
                    )
                )
                evidence_count += 1
    return {
        "source_rows": source_rows,
        "geographies": len(aggregates),
        "evidence_records": evidence_count,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Ingest verified DS057 regional evidence")
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--sqlite", type=Path, required=True)
    parser.add_argument("--maximum-source-rows", type=int)
    parser.add_argument("--expected-sha256")
    parser.add_argument("--skip-checksum", action="store_true")
    args = parser.parse_args()
    result = ingest_livestock_workbook(
        args.workbook,
        EvidenceStore(args.sqlite),
        verify_checksum=not args.skip_checksum,
        expected_sha256=args.expected_sha256 or EXPECTED_SHA256,
        maximum_source_rows=args.maximum_source_rows,
    )
    print(result)
