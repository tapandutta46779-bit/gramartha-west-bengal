from __future__ import annotations

import argparse
import hashlib
import json
from collections import Counter, defaultdict
from datetime import UTC, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

SOURCE_SHA256 = "8917232b56c61a04d72ac7a052eff962aa3a53832a587587e1dfe493b12aa595"
SOURCE_URL = (
    "https://dahd.gov.in/sites/default/files/2023-07/VillageAndWardLevelDataMale-Female.xlsx"
)
SHEETS = (
    "Rural Male Population",
    "Rural Female Population",
    "Urban Male Population",
    "Urban Female Population",
)
SPECIES = ("cattle", "buffalo", "sheep", "goat", "pig")


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def extract(source: Path, output: Path, manifest: Path) -> dict:
    actual_hash = sha256(source)
    if actual_hash != SOURCE_SHA256:
        raise ValueError(f"unexpected DS057 source SHA-256: {actual_hash}")
    source_book = load_workbook(source, read_only=True, data_only=True)
    target_book = Workbook(write_only=True)
    readme = target_book.create_sheet("README")
    readme.append(["DS057 — all available West Bengal village/ward livestock rows"])
    readme.append(["Publisher", "Department of Animal Husbandry and Dairying"])
    readme.append(["Official source", SOURCE_URL])
    readme.append(["Source SHA-256", SOURCE_SHA256])
    readme.append(["Filter", "state_name exactly equals West Bengal"])
    readme.append(["Dataset version", "20th Livestock Census, reference year 2019"])
    readme.append(["Evidence type", "OBSERVED census table row"])
    readme.append(["Note", "District/locality spellings are preserved exactly from the source."])

    sheet_counts: dict[str, int] = {}
    district_counts: dict[str, Counter] = {}
    district_totals: dict[str, dict[str, Counter]] = defaultdict(lambda: defaultdict(Counter))
    for sheet_name in SHEETS:
        source_sheet = source_book[sheet_name]
        rows = source_sheet.iter_rows(values_only=True)
        next(rows)
        next(rows)
        header = tuple(next(rows))
        positions = {str(name): index for index, name in enumerate(header)}
        target_sheet = target_book.create_sheet(sheet_name)
        target_sheet.append((*header, "source_row"))
        counts = Counter()
        retained = 0
        for source_row, row in enumerate(rows, start=4):
            if row[positions["state_name"]] != "West Bengal":
                continue
            district = str(row[positions["district_name"]])
            target_sheet.append((*row, source_row))
            retained += 1
            counts[district] += 1
            for species in SPECIES:
                district_totals[district][sheet_name][species] += int(row[positions[species]] or 0)
        sheet_counts[sheet_name] = retained
        district_counts[sheet_name] = counts
    output.parent.mkdir(parents=True, exist_ok=True)
    target_book.save(output)
    output_hash = sha256(output)
    all_districts = sorted({district for counts in district_counts.values() for district in counts})
    result = {
        "dataset_id": "DS057-WB-ALL",
        "dataset_version": "20th Livestock Census, reference year 2019",
        "created_at": datetime.now(UTC).isoformat(),
        "source_url": SOURCE_URL,
        "source_filename": source.name,
        "source_size_bytes": source.stat().st_size,
        "source_sha256": actual_hash,
        "publisher_subset_availability": (
            "No West Bengal-only publisher workbook was found; national source was required."
        ),
        "filter": "state_name exactly equals West Bengal",
        "output_filename": output.name,
        "output_size_bytes": output.stat().st_size,
        "output_sha256": output_hash,
        "districts": all_districts,
        "district_count": len(all_districts),
        "sheet_rows": sheet_counts,
        "district_rows_by_sheet": {
            sheet: dict(sorted(counts.items())) for sheet, counts in district_counts.items()
        },
        "district_species_totals_by_sheet": {
            district: {
                sheet: dict(species_counts) for sheet, species_counts in sorted(sheet_data.items())
            }
            for district, sheet_data in sorted(district_totals.items())
        },
        "evidence_type": "OBSERVED",
        "source_row_preserved": True,
        "missing_or_unavailable": [],
    }
    manifest.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    return result


def main() -> None:
    parser = argparse.ArgumentParser(description="Extract every West Bengal DS057 source row")
    parser.add_argument("source", type=Path)
    parser.add_argument("output", type=Path)
    parser.add_argument("--manifest", type=Path, required=True)
    args = parser.parse_args()
    print(json.dumps(extract(args.source, args.output, args.manifest), indent=2))


if __name__ == "__main__":
    main()
