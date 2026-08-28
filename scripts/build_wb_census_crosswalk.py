from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import unicodedata
from collections import Counter, defaultdict
from datetime import UTC, datetime
from pathlib import Path

from openpyxl import load_workbook

from backend.evidence.store import EvidenceStore
from backend.models.geography import GeographicIdentity


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def normalize(value: str) -> str:
    value = unicodedata.normalize("NFKD", value).casefold()
    value = re.sub(r"\([^)]*\)", " ", value)
    return " ".join(re.sub(r"[^a-z0-9\s]", " ", value).split())


DISTRICT_ALIASES = {
    "darjiling": {"darjeeling"},
    "koch bihar": {"coochbehar", "cooch behar"},
    "uttar dinajpur": {"dinajpur uttar"},
    "dakshin dinajpur": {"dinajpur dakshin"},
    "maldah": {"maldah", "malda"},
    "north twenty four parganas": {"24 paraganas north", "north 24 parganas"},
    "hugli": {"hooghly"},
    "puruliya": {"purulia"},
    "haora": {"howrah"},
    "south twenty four parganas": {"24 paraganas south", "south 24 parganas"},
    "paschim medinipur": {"medinipur west"},
    "purba medinipur": {"medinipur east"},
}


def district_keys(name: str) -> set[str]:
    canonical = normalize(name)
    keys = {canonical, *DISTRICT_ALIASES.get(canonical, set())}
    for official, aliases in DISTRICT_ALIASES.items():
        if canonical in aliases:
            keys.update({official, *aliases})
    return keys


def extract_west_bengal(source: Path, output_csv: Path) -> list[dict[str, str]]:
    sheet = load_workbook(source, read_only=True, data_only=True).active
    raw_rows = []
    district_names = {}
    subdistrict_names = {}
    for values in sheet.iter_rows(min_row=2, values_only=True):
        state, district, subdistrict, locality_code, name = (str(value).strip() for value in values)
        if state.zfill(2) != "19":
            continue
        district = district.zfill(3)
        subdistrict = subdistrict.zfill(5)
        locality_code = locality_code.zfill(6)
        raw_rows.append((district, subdistrict, locality_code, name))
        if district != "000" and subdistrict == "00000" and locality_code == "000000":
            district_names[district] = name
        elif district != "000" and subdistrict != "00000" and locality_code == "000000":
            subdistrict_names[(district, subdistrict)] = name
    records = []
    for district, subdistrict, locality_code, name in raw_rows:
        if locality_code == "000000":
            continue
        records.append(
            {
                "state_code": "19",
                "district_code": district,
                "district": district_names.get(district, ""),
                "subdistrict_code": subdistrict,
                "subdistrict": subdistrict_names.get((district, subdistrict), ""),
                "town_village_code": locality_code,
                "town_village_name": name,
                "locality_type": "TOWN" if int(locality_code) >= 800000 else "VILLAGE",
            }
        )
    output_csv.parent.mkdir(parents=True, exist_ok=True)
    with output_csv.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(records[0]))
        writer.writeheader()
        writer.writerows(records)
    return records


def reconcile(records: list[dict[str, str]], database: Path) -> dict[str, object]:
    store = EvidenceStore(database)
    by_district_locality: dict[tuple[str, str], list[dict[str, str]]] = defaultdict(list)
    for record in records:
        for district_key in district_keys(record["district"]):
            by_district_locality[(district_key, normalize(record["town_village_name"]))].append(
                record
            )
    matched_codes: set[str] = set()
    methods = Counter()
    ambiguous = 0
    with store.transaction():
        for geography in store.all_geographies():
            candidates = []
            for district_key in district_keys(geography.district):
                candidates.extend(
                    by_district_locality.get((district_key, normalize(geography.locality)), [])
                )
            candidates = list({item["town_village_code"]: item for item in candidates}.values())
            method = "EXACT_DISTRICT_LOCALITY"
            if len(candidates) > 1 and geography.block:
                block = normalize(geography.block)
                narrowed = [item for item in candidates if normalize(item["subdistrict"]) == block]
                if len(narrowed) == 1:
                    candidates = narrowed
                    method = "EXACT_DISTRICT_SUBDISTRICT_LOCALITY"
            if len(candidates) == 1:
                candidate = candidates[0]
                matched_codes.add(candidate["town_village_code"])
                store.put_geography(
                    geography.model_copy(
                        update={
                            "census_code": candidate["town_village_code"],
                            "source_ids": sorted({*geography.source_ids, "CENSUS2011-LCD"}),
                            "quality_flags": sorted(
                                {
                                    *geography.quality_flags,
                                    "CENSUS_2011_CODE_EXACT_NAME_HIERARCHY",
                                    method,
                                }
                            ),
                        }
                    )
                )
                methods[method] += 1
            elif len(candidates) > 1:
                ambiguous += 1
        added = 0
        added_codes: set[str] = set()
        for record in records:
            code = record["town_village_code"]
            if code in matched_codes or code in added_codes:
                continue
            district = record["district"] or "West Bengal statewide urban listing"
            locality = record["town_village_name"]
            store.put_geography(
                GeographicIdentity(
                    geo_id=f"CENSUS2011:19:{code}",
                    district=district,
                    locality=locality,
                    locality_type=record["locality_type"],
                    census_code=code,
                    block=record["subdistrict"] or None,
                    aliases=[normalize(locality)]
                    if normalize(locality) != locality.casefold()
                    else [],
                    source_ids=["CENSUS2011-LCD"],
                    quality_flags=[
                        "OFFICIAL_CENSUS_2011_LOCATION_CODE",
                        "OBSERVATION_YEAR_2011",
                        "CURRENT_ADMIN_ALIGNMENT_NOT_VERIFIED",
                    ],
                )
            )
            added_codes.add(code)
            added += 1
    return {
        "ds057_geographies_matched": sum(methods.values()),
        "match_methods": dict(methods),
        "ambiguous_ds057_not_merged": ambiguous,
        "official_census_geographies_added": added,
        "official_records": len(records),
        "official_unique_location_codes": len({record["town_village_code"] for record in records}),
        "duplicate_location_code_rows": len(records)
        - len({record["town_village_code"] for record in records}),
    }


def build(source: Path, output_csv: Path, database: Path, manifest_path: Path) -> dict:
    records = extract_west_bengal(source, output_csv)
    result = reconcile(records, database)
    result.update(
        {
            "dataset_id": "CENSUS2011-PC11-TV-DIR-WB",
            "dataset_version": "Census 2011 Location Code Directory",
            "official_source_url": ("https://censusindia.gov.in/nada/index.php/catalog/42648"),
            "national_source_required_reason": (
                "Publisher provides one national directory; state code 19 was filtered."
            ),
            "source_size_bytes": source.stat().st_size,
            "source_sha256": sha256(source),
            "west_bengal_csv_size_bytes": output_csv.stat().st_size,
            "west_bengal_csv_sha256": sha256(output_csv),
            "created_at": datetime.now(UTC).isoformat(),
            "limitations": [
                "Codes and administrative hierarchy describe Census 2011, not current LGD.",
                "Post-2011 district splits require a separate current LGD reconciliation.",
                "Ambiguous names are not silently merged.",
            ],
        }
    )
    manifest_path.parent.mkdir(parents=True, exist_ok=True)
    manifest_path.write_text(json.dumps(result, indent=2, ensure_ascii=False) + "\n")
    return result


def main() -> None:
    parser = argparse.ArgumentParser(description="Build official Census 2011 WB crosswalk")
    parser.add_argument("source", type=Path)
    parser.add_argument("output_csv", type=Path)
    parser.add_argument("--sqlite", type=Path, required=True)
    parser.add_argument("--manifest", type=Path, required=True)
    args = parser.parse_args()
    print(json.dumps(build(args.source, args.output_csv, args.sqlite, args.manifest), indent=2))


if __name__ == "__main__":
    main()
