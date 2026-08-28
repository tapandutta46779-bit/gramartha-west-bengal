from __future__ import annotations

import argparse
import json
from datetime import UTC, date, datetime
from pathlib import Path

from openpyxl import load_workbook

from backend.evidence.store import EvidenceStore
from backend.models.evidence import ConfidenceLevel, EvidenceRecord, EvidenceType
from backend.models.geography import GeographicIdentity


def ingest(input_dir: Path, source_manifest: Path, database: Path) -> dict:
    manifest = json.loads(source_manifest.read_text())
    source_by_filename = {item["filename"]: item for item in manifest["files"]}
    store = EvidenceStore(database)
    geography_by_code = {}
    for geography in store.all_geographies():
        if geography.census_code:
            geography_by_code.setdefault(geography.census_code, geography)
    counts = {"source_rows": 0, "geography_upserts": 0, "evidence_upserts": 0}
    ward_geo_ids: set[str] = set()
    evidence_ids: set[str] = set()
    levels = {}
    with store.transaction():
        for path in sorted(input_dir.glob("*.xlsx")):
            source = source_by_filename.get(path.name)
            if source is None:
                raise ValueError(f"{path.name} is absent from the verified download manifest")
            workbook = load_workbook(path, read_only=True, data_only=True)
            sheet = workbook.active
            rows = sheet.iter_rows(values_only=True)
            headers = [str(value) for value in next(rows)]
            position = {name: index for index, name in enumerate(headers)}
            for source_row, row in enumerate(rows, start=2):
                level = str(row[position["Level"]] or "").strip().upper()
                tru = str(row[position["TRU"]] or "").strip().upper()
                levels[level] = levels.get(level, 0) + 1
                if level not in {"VILLAGE", "TOWN", "WARD"}:
                    continue
                if tru not in {"RURAL", "URBAN"}:
                    continue
                locality_code = str(row[position["Town/Village"]] or "").zfill(6)
                ward_code = str(row[position["Ward"]] or "").zfill(4)
                if locality_code == "000000":
                    continue
                geography = geography_by_code.get(locality_code)
                if level == "WARD" or ward_code != "0000":
                    geo_id = f"CENSUS2011:19:{locality_code}:{ward_code}"
                    geography = store.get_geography(geo_id)
                    if geography is None:
                        parent = geography_by_code.get(locality_code)
                        geography = GeographicIdentity(
                            geo_id=geo_id,
                            district=parent.district if parent else str(row[position["District"]]),
                            locality=str(row[position["Name"]]).strip(),
                            locality_type="WARD",
                            census_code=f"{locality_code}:{ward_code}",
                            ward=ward_code,
                            municipality=parent.locality if parent else None,
                            source_ids=["CENSUS2011-PCA-TV"],
                            quality_flags=[
                                "OFFICIAL_CENSUS_2011_WARD_CODE",
                                "OBSERVATION_YEAR_2011",
                            ],
                        )
                        store.put_geography(geography)
                        counts["geography_upserts"] += 1
                    ward_geo_ids.add(geo_id)
                if geography is None:
                    continue
                counts["source_rows"] += 1
                for variable, column, unit in (
                    ("households_observed_2011", "No_HH", "households"),
                    ("population_observed_2011", "TOT_P", "persons"),
                    ("population_male_observed_2011", "TOT_M", "persons"),
                    ("population_female_observed_2011", "TOT_F", "persons"),
                ):
                    evidence_id = f"PCA2011:{geography.geo_id}:{variable}"
                    evidence_ids.add(evidence_id)
                    store.put_evidence(
                        EvidenceRecord(
                            id=evidence_id,
                            variable=variable,
                            value=int(row[position[column]] or 0),
                            unit=unit,
                            geography=geography.locality,
                            geo_id=geography.geo_id,
                            source_id="CENSUS2011-PCA-TV",
                            source_url=source["catalog_url"],
                            source_dataset="Census 2011 Primary Census Abstract TV",
                            observation_date=date(2011, 3, 1),
                            retrieved_at=datetime.fromisoformat(manifest["created_at"]),
                            evidence_type=EvidenceType.OBSERVED,
                            confidence=ConfidenceLevel.HIGH,
                            quality_flags=[
                                "OBSERVATION_YEAR_2011",
                                "NOT_CURRENT_POPULATION",
                            ],
                            methodology_version="census-pca-direct-row-v1",
                            raw_reference=path.name,
                            attributes={
                                "source_row": source_row,
                                "level": level,
                                "rural_urban": tru,
                                "dataset_version": "Census 2011",
                                "source_sha256": source["sha256"],
                            },
                        )
                    )
                    counts["evidence_upserts"] += 1
    return {
        **counts,
        "unique_ward_geographies": len(ward_geo_ids),
        "unique_evidence_records": len(evidence_ids),
        "duplicate_evidence_upserts": counts["evidence_upserts"] - len(evidence_ids),
        "source_files": len(source_by_filename),
        "level_rows_seen": levels,
        "observation_year": 2011,
        "projection_created": False,
        "warning": "No current-year projection is created or implied.",
        "completed_at": datetime.now(UTC).isoformat(),
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Ingest verified West Bengal Census PCA")
    parser.add_argument("input_dir", type=Path)
    parser.add_argument("source_manifest", type=Path)
    parser.add_argument("--sqlite", type=Path, required=True)
    parser.add_argument("--report", type=Path, required=True)
    args = parser.parse_args()
    result = ingest(args.input_dir, args.source_manifest, args.sqlite)
    args.report.parent.mkdir(parents=True, exist_ok=True)
    args.report.write_text(json.dumps(result, indent=2) + "\n")
    print(json.dumps(result, indent=2))


if __name__ == "__main__":
    main()
